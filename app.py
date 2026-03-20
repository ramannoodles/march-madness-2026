"""
app.py  —  2026 Office March Madness Web App
Single-file version: no templates/ folder needed.
Run:     python app.py
Deploy:  push to GitHub → Railway auto-deploys
"""

from flask import Flask, render_template_string, jsonify
from openpyxl import load_workbook
import os, glob
from datetime import datetime

app = Flask(__name__)

PLAYERS = ["Raman", "Porter", "Chuck", "Shelly", "Kim", "Tucker"]

ROUND_POINTS = {
    "Round of 64":  1,
    "Round of 32":  2,
    "Sweet 16":     4,
    "Elite 8":      8,
    "Final 4":      16,
    "Championship": 32,
}


def find_excel_file():
    env_file = os.environ.get("EXCEL_FILE")
    if env_file and os.path.exists(env_file):
        return env_file
    search_dir = os.path.dirname(os.path.abspath(__file__))
    for pattern in ["*Bracket*.xlsx", "*bracket*.xlsx", "*Office*.xlsx", "*.xlsx"]:
        matches = glob.glob(os.path.join(search_dir, pattern))
        if matches:
            return matches[0]
    raise FileNotFoundError("No Excel bracket file found in " + search_dir)


def load_data():
    wb = load_workbook(find_excel_file(), data_only=True)

    wm = wb["Master_Results"]
    results = {}
    for row in wm.iter_rows(min_row=1, max_row=wm.max_row, values_only=True):
        gid = row[1] if len(row) > 1 else None
        if not gid or not isinstance(gid, str):
            continue
        if not gid.startswith(("R64","Rou","Swe","Eli","Fin","Cha")):
            continue
        results[gid] = {"round": row[0], "matchup": row[2], "winner": row[3]}

    players = {}
    for p in PLAYERS:
        if p not in wb.sheetnames:
            continue
        ws = wb[p]
        picks = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            gid = row[1] if len(row) > 1 else None
            if not gid or not isinstance(gid, str):
                continue
            if not gid.startswith(("R64","Rou","Swe","Eli","Fin","Cha")):
                continue
            rnd     = row[0]
            matchup = row[2]
            pick    = row[3]
            pts_val = row[4] if row[4] else ROUND_POINTS.get(rnd, 1)
            actual  = results.get(gid, {}).get("winner")
            if actual is None:
                result = None
            elif pick and str(pick).strip().lower() == str(actual).strip().lower():
                result = 1
            else:
                result = 0
            picks.append({
                "game_id": gid, "round": rnd, "matchup": matchup,
                "pick": pick, "pts_value": pts_val,
                "result": result, "score": pts_val if result == 1 else 0,
            })
        total         = sum(pk["score"] for pk in picks)
        pts_remaining = sum(pk["pts_value"] for pk in picks if pk["result"] is None)
        players[p]    = {"picks": picks, "total": total,
                         "pts_remaining": pts_remaining,
                         "max_possible": total + pts_remaining}

    leaderboard = sorted(
        [{"name": n, **d} for n, d in players.items()],
        key=lambda x: (-x["total"], x["name"])
    )
    prev, rank_ctr = None, 1
    for i, e in enumerate(leaderboard):
        if e["total"] != prev: rank_ctr = i + 1
        e["rank"] = rank_ctr
        prev = e["total"]
        e["champion_pick"] = next(
            (pk["pick"] for pk in players[e["name"]]["picks"] if pk["game_id"] == "Cha_G1"), "—"
        )

    completed   = sum(1 for v in results.values() if v["winner"])
    total_games = len(results)
    return {
        "leaderboard": leaderboard, "players": players, "results": results,
        "completed_games": completed, "total_games": total_games,
        "last_updated": datetime.now().strftime("%b %d, %Y %I:%M %p"),
    }


HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0"/>
<title>🏀 2026 Office March Madness</title>
<style>
:root{--navy:#1a3b6f;--orange:#e8622a;--gold:#f5a623;--bg:#0f1923;--card:#1c2a3a;--card2:#243444;--text:#e8edf2;--muted:#8899aa;--green:#2ecc71;--red:#e74c3c;--border:rgba(255,255,255,0.07);}
*{box-sizing:border-box;margin:0;padding:0;}
body{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;background:var(--bg);color:var(--text);min-height:100vh;}
header{background:linear-gradient(135deg,var(--navy),#0d2548);padding:20px 16px 16px;text-align:center;border-bottom:3px solid var(--orange);position:sticky;top:0;z-index:100;box-shadow:0 4px 20px rgba(0,0,0,.4);}
header h1{font-size:clamp(17px,5vw,24px);font-weight:800;}
header h1 span{color:var(--orange);}
.header-meta{font-size:11px;color:var(--muted);margin-top:4px;}
.prog-wrap{margin:10px auto 0;max-width:360px;background:rgba(255,255,255,.1);border-radius:20px;height:8px;overflow:hidden;}
.prog-fill{background:linear-gradient(90deg,var(--orange),var(--gold));height:100%;border-radius:20px;transition:width .6s ease;}
nav{display:flex;overflow-x:auto;background:var(--card);border-bottom:1px solid var(--border);scrollbar-width:none;}
nav::-webkit-scrollbar{display:none;}
nav a{flex-shrink:0;padding:12px 18px;color:var(--muted);text-decoration:none;font-size:13px;font-weight:600;border-bottom:3px solid transparent;white-space:nowrap;cursor:pointer;transition:all .2s;}
nav a.active{color:var(--orange);border-bottom-color:var(--orange);}
main{max-width:700px;margin:0 auto;padding:16px;}
section{display:none;}
section.active{display:block;}

/* Leaderboard */
.lb-card{background:var(--card);border-radius:14px;overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,.3);margin-bottom:16px;}
.lb-hdr{display:grid;grid-template-columns:44px 1fr 80px 70px;padding:8px 16px;background:var(--navy);font-size:10px;text-transform:uppercase;letter-spacing:1px;color:var(--muted);}
.lb-row{display:grid;grid-template-columns:44px 1fr 80px 70px;align-items:center;padding:14px 16px;border-bottom:1px solid var(--border);cursor:pointer;transition:background .2s;}
.lb-row:last-child{border-bottom:none;}
.lb-row:hover{background:var(--card2);}
.r1{background:linear-gradient(135deg,rgba(245,166,35,.15),rgba(245,166,35,.05));}
.r2{background:linear-gradient(135deg,rgba(180,180,180,.1),rgba(180,180,180,.03));}
.r3{background:linear-gradient(135deg,rgba(205,127,50,.1),rgba(205,127,50,.03));}
.badge{width:32px;height:32px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:18px;}
.lb-name{font-weight:700;font-size:16px;}
.lb-champ{font-size:11px;color:var(--muted);margin-top:2px;}
.lb-champ::before{content:"🏆 ";}
.lb-score{text-align:right;}
.lb-score-num{font-size:22px;font-weight:800;color:var(--orange);}
.lb-score-lbl{font-size:10px;color:var(--muted);}
.lb-max{text-align:right;}
.lb-max-num{font-size:13px;font-weight:600;}
.lb-max-lbl{font-size:10px;color:var(--muted);}
.lb-max-rem{font-size:11px;color:var(--green);}

/* Player tabs */
.ptabs{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:16px;}
.ptab{padding:8px 16px;border-radius:20px;background:var(--card);color:var(--muted);font-size:13px;font-weight:600;cursor:pointer;border:2px solid transparent;transition:all .2s;}
.ptab.active{border-color:var(--orange);color:var(--orange);background:rgba(232,98,42,.1);}

/* Score summary */
.ssumm{display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px;margin-bottom:16px;}
.scard{background:var(--card);border-radius:12px;padding:14px;text-align:center;}
.scard-num{font-size:28px;font-weight:800;color:var(--orange);}
.scard-lbl{font-size:11px;color:var(--muted);margin-top:2px;}

/* Bracket rows */
.round-hdr{background:linear-gradient(135deg,var(--navy),#0d2548);padding:10px 14px;border-radius:10px 10px 0 0;font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:1px;color:var(--gold);}
.game-row{display:grid;grid-template-columns:1fr auto auto;align-items:center;padding:10px 14px;background:var(--card);border-bottom:1px solid var(--border);gap:8px;}
.game-row:last-child{border-bottom:none;border-radius:0 0 10px 10px;}
.game-matchup{font-size:11px;color:var(--muted);}
.game-pick{font-size:14px;font-weight:700;}
.correct{color:var(--green);}
.wrong{color:var(--red);text-decoration:line-through;}
.pending{color:var(--text);}
.pill{font-size:11px;padding:3px 8px;border-radius:20px;font-weight:700;white-space:nowrap;}
.pill-c{background:rgba(46,204,113,.15);color:var(--green);}
.pill-w{background:rgba(231,76,60,.15);color:var(--red);}
.pill-p{background:rgba(255,255,255,.07);color:var(--muted);}
.gpts{font-size:13px;font-weight:700;color:var(--orange);min-width:28px;text-align:right;}

/* Results */
.res-row{display:grid;grid-template-columns:1fr auto;padding:10px 14px;background:var(--card);border-bottom:1px solid var(--border);gap:8px;align-items:center;}
.res-row:last-child{border-bottom:none;border-radius:0 0 10px 10px;}
.res-matchup{font-size:13px;}
.res-pill-done{font-size:12px;font-weight:700;padding:4px 12px;border-radius:20px;background:rgba(46,204,113,.15);color:var(--green);}
.res-pill-pend{font-size:11px;padding:4px 10px;border-radius:20px;background:rgba(255,255,255,.06);color:var(--muted);}

/* Comparison */
.cmp-wrap{overflow-x:auto;border-radius:10px;}
.cmp{width:100%;border-collapse:collapse;font-size:12px;}
.cmp th{background:var(--navy);padding:8px 6px;text-align:center;font-size:11px;color:var(--muted);}
.cmp td{padding:8px 6px;text-align:center;border-bottom:1px solid var(--border);background:var(--card);}
.cmp tr:hover td{background:var(--card2);}
.cc{color:var(--green);font-weight:700;}
.cw{color:var(--red);}
.cp{color:var(--muted);}
.sec-row td{background:var(--navy)!important;color:var(--gold);font-weight:700;font-size:10px;text-transform:uppercase;letter-spacing:1px;padding:5px 6px;}
.mtch{text-align:left!important;color:var(--muted);font-size:11px;}
.act{color:var(--green);font-weight:700;}

.last-updated{text-align:center;font-size:11px;color:var(--muted);padding:12px;margin-top:8px;}
@media(max-width:480px){
  .lb-row{grid-template-columns:40px 1fr 70px 58px;padding:12px;}
  .lb-score-num{font-size:18px;}
  .scard-num{font-size:22px;}
}
</style>
</head>
<body>

<header>
  <h1>🏀 2026 <span>MARCH MADNESS</span> 🏀</h1>
  <div class="header-meta" id="game-count"></div>
  <div class="prog-wrap"><div class="prog-fill" id="prog-bar"></div></div>
</header>

<nav>
  <a class="active" onclick="showTab('leaderboard',this)">🏆 Leaderboard</a>
  <a onclick="showTab('brackets',this)">📋 Brackets</a>
  <a onclick="showTab('results',this)">🎯 Results</a>
  <a onclick="showTab('comparison',this)">📊 Compare</a>
</nav>

<main>
  <section id="leaderboard" class="active"></section>
  <section id="brackets"></section>
  <section id="results"></section>
  <section id="comparison"></section>
</main>

<script>
const DATA = __DATA__;
const ROUND_ORDER = ["Round of 64","Round of 32","Sweet 16","Elite 8","Final 4","Championship"];
const ROUND_LABELS = {
  "Round of 64":"Round of 64 · 1 pt each",
  "Round of 32":"Round of 32 · 2 pts each",
  "Sweet 16":"Sweet Sixteen · 4 pts each",
  "Elite 8":"Elite Eight · 8 pts each",
  "Final 4":"Final Four · 16 pts each",
  "Championship":"Championship · 32 pts"
};
const MEDALS = {1:"🥇",2:"🥈",3:"🥉"};

function showTab(id, el) {
  document.querySelectorAll("section").forEach(s=>s.classList.remove("active"));
  document.querySelectorAll("nav a").forEach(a=>a.classList.remove("active"));
  document.getElementById(id).classList.add("active");
  el.classList.add("active");
}

function selectPlayer(name, switchTab=false) {
  document.querySelectorAll(".ptab").forEach(t=>t.classList.remove("active"));
  const tab = document.getElementById("ptab-"+name);
  if(tab) tab.classList.add("active");
  document.querySelectorAll(".bracket-pane").forEach(b=>b.style.display="none");
  const pane = document.getElementById("bp-"+name);
  if(pane) pane.style.display="block";
  if(switchTab){
    const navLink = document.querySelector("nav a:nth-child(2)");
    showTab("brackets", navLink);
  }
  window.scrollTo({top:0,behavior:"smooth"});
}

function buildLeaderboard() {
  const pct = DATA.total_games > 0 ? (DATA.completed_games/DATA.total_games*100).toFixed(1) : 0;
  document.getElementById("game-count").textContent = DATA.completed_games+" / "+DATA.total_games+" games complete";
  document.getElementById("prog-bar").style.width = pct+"%";

  let html = `<div class="lb-card">
    <div class="lb-hdr"><div></div><div>Player</div><div style="text-align:right">Score</div><div style="text-align:right">Max</div></div>`;
  DATA.leaderboard.forEach(e=>{
    const cls = e.rank<=3 ? "r"+e.rank : "";
    const medal = MEDALS[e.rank] || e.rank;
    html += `<div class="lb-row ${cls}" onclick="selectPlayer('${e.name}',true)">
      <div><div class="badge">${medal}</div></div>
      <div>
        <div class="lb-name">${e.name}</div>
        <div class="lb-champ">${e.champion_pick}</div>
      </div>
      <div class="lb-score">
        <div class="lb-score-num">${e.total}</div>
        <div class="lb-score-lbl">pts</div>
      </div>
      <div class="lb-max">
        <div class="lb-max-num">${e.max_possible}</div>
        <div class="lb-max-lbl">max</div>
        <div class="lb-max-rem">+${e.pts_remaining} left</div>
      </div>
    </div>`;
  });
  html += `</div><div class="last-updated">Last updated: ${DATA.last_updated}</div>`;
  document.getElementById("leaderboard").innerHTML = html;
}

function buildBrackets() {
  const players = Object.keys(DATA.players);
  let tabsHtml = '<div class="ptabs">';
  players.forEach((p,i) => {
    tabsHtml += `<div class="ptab${i===0?' active':''}" id="ptab-${p}" onclick="selectPlayer('${p}')">${p}</div>`;
  });
  tabsHtml += "</div>";

  let panesHtml = "";
  players.forEach((p,pi) => {
    const pl = DATA.players[p];
    let pane = `<div class="bracket-pane" id="bp-${p}" style="display:${pi===0?'block':'none'}">
      <div class="ssumm">
        <div class="scard"><div class="scard-num">${pl.total}</div><div class="scard-lbl">Current</div></div>
        <div class="scard"><div class="scard-num">${pl.pts_remaining}</div><div class="scard-lbl">Remaining</div></div>
        <div class="scard"><div class="scard-num">${pl.max_possible}</div><div class="scard-lbl">Max Possible</div></div>
      </div>`;

    let curRound = null;
    pl.picks.forEach((pick, i) => {
      if(pick.round !== curRound) {
        if(curRound !== null) pane += "</div>";
        pane += `<div style="margin-bottom:20px"><div class="round-hdr">${ROUND_LABELS[pick.round]||pick.round}</div>`;
        curRound = pick.round;
      }
      const cls = pick.result===1?"correct":pick.result===0?"wrong":"pending";
      const pillCls = pick.result===1?"pill-c":pick.result===0?"pill-w":"pill-p";
      const pillTxt = pick.result===1?"✓ Correct":pick.result===0?"✗ Wrong":"Pending";
      const pts = pick.result===1?"+"+pick.pts_value:pick.result===0?"0":"—";
      pane += `<div class="game-row">
        <div>
          <div class="game-matchup">${pick.matchup||""}</div>
          <div class="game-pick ${cls}">${pick.pick||"—"}</div>
        </div>
        <span class="pill ${pillCls}">${pillTxt}</span>
        <div class="gpts">${pts}</div>
      </div>`;
    });
    if(curRound !== null) pane += "</div>";
    pane += "</div>";
    panesHtml += pane;
  });

  document.getElementById("brackets").innerHTML = tabsHtml + panesHtml;
}

function buildResults() {
  let html = "";
  ROUND_ORDER.forEach(rnd => {
    const games = Object.values(DATA.results).filter(g=>g.round===rnd);
    if(!games.length) return;
    html += `<div style="margin-bottom:20px"><div class="round-hdr">${rnd}</div>`;
    games.forEach(g => {
      html += `<div class="res-row">
        <div class="res-matchup">${g.matchup||""}</div>
        ${g.winner
          ? `<span class="res-pill-done">✓ ${g.winner}</span>`
          : `<span class="res-pill-pend">Pending</span>`}
      </div>`;
    });
    html += "</div>";
  });
  document.getElementById("results").innerHTML = html;
}

function buildComparison() {
  const players = Object.keys(DATA.players);
  let html = `<div class="cmp-wrap"><table class="cmp"><thead><tr>
    <th style="text-align:left;min-width:140px">Matchup</th>
    <th>Actual</th>
    ${players.map(p=>`<th>${p}</th>`).join("")}
  </tr></thead><tbody>`;

  let curRound = null;
  Object.entries(DATA.results).forEach(([gid, game]) => {
    if(game.round !== curRound) {
      html += `<tr class="sec-row"><td colspan="${2+players.length}">${game.round}</td></tr>`;
      curRound = game.round;
    }
    html += `<tr><td class="mtch">${game.matchup||""}</td><td class="act">${game.winner||"—"}</td>`;
    players.forEach(p => {
      const pick = DATA.players[p].picks.find(pk=>pk.game_id===gid);
      const cls = !pick ? "cp" : pick.result===1?"cc":pick.result===0?"cw":"cp";
      html += `<td class="${cls}">${pick?pick.pick||"—":"—"}</td>`;
    });
    html += "</tr>";
  });
  html += "</tbody></table></div>";
  document.getElementById("comparison").innerHTML = html;
}

// Init
buildLeaderboard();
buildBrackets();
buildResults();
buildComparison();

// Auto-refresh every 3 minutes
setInterval(()=>location.reload(), 180000);
</script>
</body>
</html>"""


@app.route("/")
def index():
    import json
    data = load_data()
    # Convert to JSON-serializable format for embedding in JS
    js_data = {
        "leaderboard": [
            {k: v for k, v in e.items() if k != "picks"} for e in data["leaderboard"]
        ],
        "players": {
            name: {
                "picks": pdata["picks"],
                "total": pdata["total"],
                "pts_remaining": pdata["pts_remaining"],
                "max_possible": pdata["max_possible"],
            }
            for name, pdata in data["players"].items()
        },
        "results": data["results"],
        "completed_games": data["completed_games"],
        "total_games": data["total_games"],
        "last_updated": data["last_updated"],
    }
    html = HTML.replace("__DATA__", json.dumps(js_data))
    return html


@app.route("/api/data")
def api_data():
    data = load_data()
    return jsonify({
        "leaderboard": [{k: v for k, v in e.items() if k != "picks"} for e in data["leaderboard"]],
        "completed_games": data["completed_games"],
        "total_games": data["total_games"],
        "last_updated": data["last_updated"],
    })


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
